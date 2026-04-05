from flask import Flask, request, send_file, render_template_string, redirect, url_for, flash
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from copy import copy
from io import BytesIO
import tempfile
import os
import re

app = Flask(__name__)
app.secret_key = "top5_secret_key"

HTML = '''
<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <title>TOP MERMAS - Generador</title>
</head>
<body style="font-family:Arial,Helvetica,sans-serif;max-width:900px;margin:20px;">
  <h2>Generador TOP (sube solo el fichero MERMAS)</h2>
  <p>Sube únicamente el fichero MERMAS (.xlsx). Las plantillas ya están integradas en la app.</p>

  <form action="/generate" method="post" enctype="multipart/form-data">
    <label>Fichero MERMAS (xlsx): <input type="file" name="mermas" accept=".xls,.xlsx,.xlsm" required></label><br><br>
    <button type="submit">Generar Excel</button>
  </form>

  <hr style="margin:28px 0;">

  <h3>Generar histórico (sube varios TOP MERMAS)</h3>
  <form action="/historico" method="post" enctype="multipart/form-data">
    <input type="file" name="files" multiple accept=".xlsx,.xlsm,.xls" required><br><br>
    <button type="submit">Generar histórico</button>
  </form>

  <p style="color:gray;font-size:0.9em">Si necesitas subir tus propias plantillas, contacta para activar la opción.</p>
</body>
</html>
'''


def copy_sheet_exact(src_ws, tgt_wb, title):
    tgt = tgt_wb.create_sheet(title=title)

    try:
        for col, dim in src_ws.column_dimensions.items():
            if getattr(dim, "width", None) is not None:
                tgt.column_dimensions[col].width = dim.width
    except Exception:
        pass

    try:
        for r, dim in src_ws.row_dimensions.items():
            if getattr(dim, "height", None) is not None:
                tgt.row_dimensions[r].height = dim.height
    except Exception:
        pass

    try:
        for merged in src_ws.merged_cells.ranges:
            tgt.merge_cells(str(merged))
    except Exception:
        pass

    max_row = src_ws.max_row
    max_col = src_ws.max_column

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            s = src_ws.cell(row=r, column=c)
            t = tgt.cell(row=r, column=c, value=s.value)
            try:
                if s.has_style:
                    t.font = copy(s.font)
                    t.border = copy(s.border)
                    t.fill = copy(s.fill)
                    t.number_format = copy(s.number_format)
                    t.protection = copy(s.protection)
                    t.alignment = copy(s.alignment)
            except Exception:
                pass
            try:
                if s.comment:
                    t.comment = copy(s.comment)
            except Exception:
                pass

    try:
        tgt.sheet_view = copy(src_ws.sheet_view)
    except Exception:
        pass

    try:
        tgt.page_setup = copy(src_ws.page_setup)
    except Exception:
        pass

    return tgt


def extract_sem_ciclo_from_name(fname):
    sem = None
    cic = None

    m = re.search(r"[sS][eE][mM][^0-9]*([0-9]{1,2})", fname)
    if m:
        sem = m.group(1)

    m2 = re.search(r"[cC]\s*[_\-]?\s*([0-9]{1,2})", fname)
    if m2:
        cic = m2.group(1)

    if not sem:
        m3 = re.search(r"\bS[_\s\-]?([0-9]{1,2})\b", fname, flags=re.IGNORECASE)
        if m3:
            sem = m3.group(1)

    if not cic:
        m4 = re.search(r"\bC[_\s\-]?([0-9]{1,2})\b", fname, flags=re.IGNORECASE)
        if m4:
            cic = m4.group(1)

    return sem, cic


def sanitize_sheet_name(name):
    s = re.sub(r'[:\\/*?\[\]]', '-', str(name))
    s = s.replace("\n", " ").strip()
    return s[:31] if len(s) > 31 else s


def convert_pct_cell_to_number(cell):
    v = cell.value
    if v is None:
        return

    try:
        if isinstance(v, (int, float)) and abs(v) < 1.0:
            cell.value = round(float(v) * 100.0, 6)
        elif isinstance(v, str) and '%' in v:
            s = v.replace('%', '').strip()
            try:
                cell.value = float(s)
            except Exception:
                pass

        try:
            cell.number_format = '0.00'
        except Exception:
            pass
    except Exception:
        pass


@app.route('/', methods=['GET'])
def index():
    return render_template_string(HTML)


@app.route('/generate', methods=['POST'])
def generate():
    if 'mermas' not in request.files:
        flash('Falta fichero MERMAS')
        return redirect(url_for('index'))

    m = request.files['mermas']
    if not m or m.filename.strip() == '':
        flash('Falta fichero MERMAS')
        return redirect(url_for('index'))

    tmpdir = tempfile.mkdtemp()
    mer_path = os.path.join(tmpdir, 'mermas.xlsx')
    m.save(mer_path)

    tpl_art = os.path.join(os.path.dirname(__file__), 'PLANTILLA Artículos.xlsx')
    tpl_chk = os.path.join(os.path.dirname(__file__), 'CHECKLIST CALIDAD DE REPARTO.xlsx')

    if not os.path.exists(tpl_art) or not os.path.exists(tpl_chk):
        return "Faltan plantillas en la app. Contacta para añadirlas.", 500

    out_path = os.path.join(tmpdir, 'TOP_GENERADO.xlsx')

    try:
        generate_from_mermas(mer_path, tpl_art, tpl_chk, out_path, m.filename)
    except Exception as e:
        return f"Error durante generación: {e}", 500

    return send_file(out_path, as_attachment=True, download_name='TOP_GENERADO.xlsx')


def generate_from_mermas(mermas_path, tpl_art_path, tpl_chk_path, out_path, original_filename):
    wb_mermas = load_workbook(mermas_path, data_only=True)

    top_name = None
    for name in wb_mermas.sheetnames:
        if 'top' in name.lower() or 'calidad' in name.lower():
            top_name = name
            break
    if not top_name:
        top_name = wb_mermas.sheetnames[0]

    ws_top = wb_mermas[top_name]

    sem, cic = extract_sem_ciclo_from_name(original_filename)
    sem = sem or ''
    cic = cic or ''

    headers = [(ws_top.cell(row=1, column=c).value or "") for c in range(1, ws_top.max_column + 1)]
    low = [str(h).lower() for h in headers]

    def find_idx(keys):
        for k in keys:
            for i, h in enumerate(low):
                if k in h:
                    return i + 1
        return None

    mc_idx = find_idx(['mc', 'modelo', 'model', 'codigo', 'articulo']) or 1
    fam_idx = find_idx(['familia', 'family', 'grupo'])

    map_cols = [2, 3, 4, 5, 6, 7, 9, 10]

    mc_map = {}
    for r in range(2, ws_top.max_row + 1):
        val = ws_top.cell(row=r, column=mc_idx).value
        if val is None:
            continue

        key = str(val).strip()
        if not key:
            continue

        mapped = []
        for col in map_cols:
            mapped.append(ws_top.cell(row=r, column=col).value if col <= ws_top.max_column else None)

        fam = ws_top.cell(row=r, column=fam_idx).value if fam_idx and fam_idx <= ws_top.max_column else None
        mc_map[key] = {'mapped': mapped, 'fam': fam, 'row': r}

    wb_art = load_workbook(tpl_art_path, data_only=False)
    wb_chk = load_workbook(tpl_chk_path, data_only=False)

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    copy_sheet_exact(ws_top, wb_out, "TOP-CALIDAD-2")

    tpl_art_sheet = wb_art[wb_art.sheetnames[0]]
    tpl_chk_sheet = wb_chk[wb_chk.sheetnames[0]]

    existing = set(wb_out.sheetnames)
    pairs = []

    for key in mc_map:
        safe = sanitize_sheet_name(key)
        base = safe
        i = 1
        while safe in existing:
            suf = f'-{i}'
            safe = base[:31 - len(suf)] + suf
            i += 1

        existing.add(safe)

        new_mc = copy_sheet_exact(tpl_art_sheet, wb_out, safe)

        new_mc.cell(row=4, column=2).value = key
        if mc_map[key]['fam'] is not None:
            new_mc.cell(row=5, column=2).value = mc_map[key]['fam']

        try:
            new_mc.cell(row=2, column=2).value = int(sem) if sem != '' else None
        except Exception:
            new_mc.cell(row=2, column=2).value = sem

        try:
            new_mc.cell(row=2, column=4).value = int(cic) if cic != '' else None
        except Exception:
            new_mc.cell(row=2, column=4).value = cic

        for idx, val in enumerate(mc_map[key]['mapped'], start=1):
            new_mc.cell(row=48, column=idx).value = val
        new_mc.cell(row=48, column=9).value = None

        cell = new_mc.cell(row=48, column=6)
        convert_pct_cell_to_number(cell)

        chk_name = f'CHECKLIST-{safe}'
        chk = copy_sheet_exact(tpl_chk_sheet, wb_out, chk_name)

        a1 = chk.cell(row=1, column=1).value or ""
        if re.search(r"art[ií]culo\s*:", str(a1), flags=re.IGNORECASE):
            new_a1 = re.sub(
                r"(art[ií]culo\s*:\s*).*",
                lambda m: m.group(1) + str(key),
                str(a1),
                flags=re.IGNORECASE,
            )
        else:
            new_a1 = str(a1).strip() + " Artículo: " + str(key)
        chk.cell(row=1, column=1).value = new_a1

        chk.cell(row=1, column=4).value = f"SEMANA: {sem}   CICLO: {cic}"

        chk.cell(row=4, column=2).value = None
        chk.cell(row=5, column=2).value = None
        for c in range(1, 9):
            chk.cell(row=48, column=c).value = None

        pairs.append((safe, chk_name))

    ordered = ['TOP-CALIDAD-2']
    for a, b in pairs:
        ordered.append(a)
        ordered.append(b)

    for s in wb_out.sheetnames:
        if s not in ordered:
            ordered.append(s)

    wb_out._sheets = [wb_out[s] for s in ordered]
    wb_out.save(out_path)


@app.route('/historico', methods=['POST'])
def generar_historico():
    uploaded_files = request.files.getlist('files')

    if not uploaded_files or all((not f or f.filename == '') for f in uploaded_files):
        return "No has subido ficheros para el histórico", 400

    valid_files = [f for f in uploaded_files if f and f.filename and f.filename.lower().endswith(('.xlsx', '.xlsm', '.xls'))]
    if not valid_files:
        return "No has subido ficheros Excel válidos para el histórico", 400

    records = {}
    last_label = None
    last_seen_models = set()

    for idx, f in enumerate(valid_files):
        try:
            file_bytes = BytesIO(f.read())
            wb = load_workbook(file_bytes, data_only=True)

            top_name = None
            for name in wb.sheetnames:
                if 'top' in name.lower() or 'calidad' in name.lower():
                    top_name = name
                    break
            if not top_name:
                top_name = wb.sheetnames[0]

            ws = wb[top_name]

            sem, cic = extract_sem_ciclo_from_name(f.filename)
            if not sem or not cic:
                top_sem = ws["B2"].value
                top_cic = ws["D2"].value
                if not sem:
                    sem_match = re.search(r'\d+', str(top_sem) if top_sem is not None else "")
                    if sem_match:
                        sem = sem_match.group(0)
                if not cic:
                    cic_match = re.search(r'\d+', str(top_cic) if top_cic is not None else "")
                    if cic_match:
                        cic = cic_match.group(0)

            label = f"SEM {sem} C{cic}" if sem or cic else f.filename

            if idx == len(valid_files) - 1:
                last_label = label

            seen_in_this_file = set()
            for r in range(2, ws.max_row + 1):
                mc = ws.cell(row=r, column=1).value
                if mc is None:
                    continue
                mc = str(mc).strip()
                if not mc:
                    continue
                seen_in_this_file.add(mc)

            for mc in seen_in_this_file:
                if mc not in records:
                    records[mc] = {
                        "count": 0,
                        "ciclos": [],
                        "first_idx": idx,
                        "first_label": label,
                        "last_idx": idx,
                        "last_label": label,
                    }

                rec = records[mc]
                rec["count"] += 1
                if label not in rec["ciclos"]:
                    rec["ciclos"].append(label)

                if idx < rec["first_idx"]:
                    rec["first_idx"] = idx
                    rec["first_label"] = label

                if idx >= rec["last_idx"]:
                    rec["last_idx"] = idx
                    rec["last_label"] = label

                if idx == len(valid_files) - 1:
                    last_seen_models.add(mc)

        except Exception:
            continue

    if not records:
        return "No se han podido leer datos válidos de los ficheros subidos", 400

    rows = []
    for mc, rec in records.items():
        appears_in_last_top = mc in last_seen_models
        is_recurrent = rec["count"] > 1
        if appears_in_last_top and is_recurrent:
            priority = "ALTA"
        elif appears_in_last_top:
            priority = "MEDIA"
        else:
            priority = "BAJA"

        rows.append({
            "mc": mc,
            "count": rec["count"],
            "ciclos": ", ".join(rec["ciclos"]),
            "first_label": rec["first_label"],
            "last_label": rec["last_label"],
            "appears_in_last_top": "Sí" if appears_in_last_top else "No",
            "estado": "Reincidente" if (appears_in_last_top and is_recurrent) else ("Nuevo" if appears_in_last_top else "Histórico"),
            "prioridad": priority,
        })

    total_models = len(rows)
    new_models = sum(1 for r in rows if r["estado"] == "Nuevo")
    repeated_models = sum(1 for r in rows if r["count"] > 1)

    priority_rank = {"ALTA": 0, "MEDIA": 1, "BAJA": 2}
    estado_rank = {"Reincidente": 0, "Nuevo": 1, "Histórico": 2}

    rows.sort(
        key=lambda r: (
            priority_rank.get(r["prioridad"], 99),
            estado_rank.get(r["estado"], 99),
            -r["count"],
            r["first_label"],
            r["mc"],
        )
    )

    wb_out = Workbook()

    ws_res = wb_out.active
    ws_res.title = "RESUMEN"
    ws_res.sheet_view.showGridLines = False

    title_font = Font(bold=True, size=14, color="FFFFFF")
    header_font = Font(bold=True, color="FFFFFF")
    body_font = Font(bold=False, color="1F1F1F")
    bold_font = Font(bold=True, color="1F1F1F")

    fill_title = PatternFill("solid", fgColor="1F4E78")
    fill_header = PatternFill("solid", fgColor="4F81BD")
    fill_card = PatternFill("solid", fgColor="D9EAF7")
    fill_green = PatternFill("solid", fgColor="E2F0D9")
    fill_amber = PatternFill("solid", fgColor="FFF2CC")
    fill_red = PatternFill("solid", fgColor="F4CCCC")
    fill_gray = PatternFill("solid", fgColor="E7E6E6")

    thin = Side(style="thin", color="7F7F7F")
    medium = Side(style="medium", color="404040")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_medium = Border(left=medium, right=medium, top=medium, bottom=medium)

    ws_res.merge_cells("A1:D1")
    ws_res["A1"] = "RESUMEN HISTÓRICO TOP MERMAS"
    ws_res["A1"].font = title_font
    ws_res["A1"].fill = fill_title
    ws_res["A1"].alignment = Alignment(horizontal="center", vertical="center")

    res_rows = [
        ("Total de modelos", total_models, fill_card),
        ("Cuántos son nuevos", new_models, fill_green if new_models else fill_gray),
        ("Cuántos se repiten", repeated_models, fill_amber if repeated_models else fill_gray),
    ]

    for i, (label, value, fill) in enumerate(res_rows, start=3):
        ws_res.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
        ws_res.merge_cells(start_row=i, start_column=3, end_row=i, end_column=4)
        c1 = ws_res.cell(row=i, column=1, value=label)
        c2 = ws_res.cell(row=i, column=3, value=value)
        c1.font = bold_font
        c2.font = bold_font
        c1.fill = fill
        c2.fill = fill
        c1.border = border_medium
        c2.border = border_medium
        c1.alignment = Alignment(horizontal="left", vertical="center")
        c2.alignment = Alignment(horizontal="center", vertical="center")

    top_start = 8
    ws_res.merge_cells(start_row=top_start, start_column=1, end_row=top_start, end_column=5)
    ws_res.cell(row=top_start, column=1, value="TOP 10 REINCIDENTES").font = Font(bold=True, color="FFFFFF")
    ws_res.cell(row=top_start, column=1).fill = fill_title
    ws_res.cell(row=top_start, column=1).alignment = Alignment(horizontal="center", vertical="center")

    res_headers = ["MC", "Veces", "Ciclos", "Estado", "Prioridad"]
    for col, h in enumerate(res_headers, start=1):
        cell = ws_res.cell(row=top_start + 1, column=col, value=h)
        cell.font = header_font
        cell.fill = fill_header
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="center", vertical="center")

    top_10 = sorted([r for r in rows if r["count"] > 1], key=lambda x: (-x["count"], x["mc"]))[:10]
    for i, r in enumerate(top_10, start=top_start + 2):
        vals = [r["mc"], r["count"], r["ciclos"], r["estado"], r["prioridad"]]
        for col, v in enumerate(vals, start=1):
            cell = ws_res.cell(row=i, column=col, value=v)
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left" if col in (1, 3, 4, 5) else "center", vertical="center")
            cell.font = body_font
            if r["prioridad"] == "ALTA":
                cell.fill = fill_red
            elif r["prioridad"] == "MEDIA":
                cell.fill = fill_amber
            else:
                cell.fill = fill_green if r["estado"] == "Nuevo" else fill_gray

    for col, width in {1: 22, 2: 12, 3: 55, 4: 16, 5: 14}.items():
        ws_res.column_dimensions[get_column_letter(col)].width = width

    ws_res.freeze_panes = "A9"

    ws_det = wb_out.create_sheet("DETALLE")
    ws_det.sheet_view.showGridLines = False

    det_headers = [
        "MC",
        "Veces",
        "Ciclos",
        "Primera aparición",
        "Última aparición",
        "¿Aparece en el último TOP?",
        "Estado",
        "Prioridad",
    ]
    for col, h in enumerate(det_headers, start=1):
        cell = ws_det.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = fill_header
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, r in enumerate(rows, start=2):
        vals = [
            r["mc"],
            r["count"],
            r["ciclos"],
            r["first_label"],
            r["last_label"],
            r["appears_in_last_top"],
            r["estado"],
            r["prioridad"],
        ]
        for col, v in enumerate(vals, start=1):
            cell = ws_det.cell(row=row_idx, column=col, value=v)
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left" if col in (1, 3, 4, 5, 6, 7, 8) else "center", vertical="center")
            cell.font = body_font
            if r["prioridad"] == "ALTA":
                cell.fill = fill_red
            elif r["prioridad"] == "MEDIA":
                cell.fill = fill_amber
            else:
                cell.fill = fill_green if r["estado"] == "Nuevo" else fill_gray

    for col, width in {
        1: 18,
        2: 10,
        3: 42,
        4: 18,
        5: 18,
        6: 24,
        7: 16,
        8: 14,
    }.items():
        ws_det.column_dimensions[get_column_letter(col)].width = width

    ws_det.freeze_panes = "A2"
    ws_det.auto_filter.ref = ws_det.dimensions

    output = BytesIO()
    wb_out.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="historico.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
