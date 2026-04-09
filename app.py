from flask import Flask, request, send_file, render_template_string, redirect, url_for, flash
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from copy import copy
from io import BytesIO
import tempfile
import os
import re

app = Flask(__name__)
app.secret_key = "top5_secret_key"

HTML = """
<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <title>TOP MERMAS - Generador</title>
</head>
<body style="font-family:Arial,Helvetica,sans-serif;max-width:900px;margin:20px;">
  <h2>Generador TOP</h2>
  <p>Sube un fichero MERMAS (.xlsx) y la app generará el TOP final.</p>

  <form action="/generate" method="post" enctype="multipart/form-data">
    <label>Fichero MERMAS (xlsx): <input type="file" name="mermas" accept=".xls,.xlsx,.xlsm" required></label><br><br>
    <button type="submit">Generar Excel</button>
  </form>

  <hr style="margin:28px 0;">

  <h2>Histórico acumulado</h2>
  <p style="color:gray;font-size:0.95em">
    Puedes subir un histórico acumulado anterior (opcional) y uno o varios TOP nuevos para actualizarlo.
  </p>

  <form action="/historico_acumulado" method="post" enctype="multipart/form-data">
    <label>Histórico acumulado anterior (opcional): <input type="file" name="master" accept=".xlsx"></label><br><br>
    <label>TOP MERMAS nuevos: <input type="file" name="files" multiple accept=".xlsx,.xlsm,.xls" required></label><br><br>
    <button type="submit">Actualizar histórico acumulado</button>
  </form>

  <hr style="margin:28px 0;">

  <h2>Panel tipo dashboard</h2>
  <p style="color:gray;font-size:0.95em">
    Sube uno o varios TOP nuevos para ver el resumen interactivo en pantalla.
  </p>

  <form action="/dashboard" method="post" enctype="multipart/form-data">
    <label>TOP MERMAS nuevos: <input type="file" name="files" multiple accept=".xlsx,.xlsm,.xls" required></label><br><br>
    <button type="submit">Ver dashboard</button>
  </form>

</body>
</html>
"""


def copy_sheet_exact(src_ws, tgt_wb, title):
    tgt = tgt_wb.create_sheet(title=title)
    try:
        for col, dim in src_ws.column_dimensions.items():
            if getattr(dim, "width", None) is not None:
                tgt.column_dimensions[col].width = dim.width
    except Exception:
        pass
    try:
        for r, dim in src_ws.row_dimensions.items():
            if getattr(dim, "height", None) is not None:
                tgt.row_dimensions[r].height = dim.height
    except Exception:
        pass
    try:
        for merged in src_ws.merged_cells.ranges:
            tgt.merge_cells(str(merged))
    except Exception:
        pass

    for r in range(1, src_ws.max_row + 1):
        for c in range(1, src_ws.max_column + 1):
            s = src_ws.cell(row=r, column=c)
            t = tgt.cell(row=r, column=c, value=s.value)
            try:
                if s.has_style:
                    t.font = copy(s.font)
                    t.border = copy(s.border)
                    t.fill = copy(s.fill)
                    t.number_format = copy(s.number_format)
                    t.protection = copy(s.protection)
                    t.alignment = copy(s.alignment)
            except Exception:
                pass
            try:
                if s.comment:
                    t.comment = copy(s.comment)
            except Exception:
                pass

    try:
        tgt.sheet_view = copy(src_ws.sheet_view)
    except Exception:
        pass
    try:
        tgt.page_setup = copy(src_ws.page_setup)
    except Exception:
        pass
    return tgt


def extract_sem_ciclo_from_name(fname):
    sem = None
    cic = None

    m = re.search(r"[sS][eE][mM][^0-9]*([0-9]{1,2})", fname)
    if m:
        sem = m.group(1)

    m2 = re.search(r"[cC]\s*[_\-]?\s*([0-9]{1,2})", fname)
    if m2:
        cic = m2.group(1)

    if not sem:
        m3 = re.search(r"\bS[_\s\-]?([0-9]{1,2})\b", fname, flags=re.IGNORECASE)
        if m3:
            sem = m3.group(1)

    if not cic:
        m4 = re.search(r"\bC[_\s\-]?([0-9]{1,2})\b", fname, flags=re.IGNORECASE)
        if m4:
            cic = m4.group(1)

    return sem, cic


def sanitize_sheet_name(name):
    s = re.sub(r'[:\\/*?\[\]]', '-', str(name))
    s = s.replace("\n", " ").strip()
    return s[:31] if len(s) > 31 else s


def convert_pct_cell_to_number(cell):
    v = cell.value
    if v is None:
        return
    try:
        if isinstance(v, (int, float)) and abs(v) < 1.0:
            cell.value = round(float(v) * 100.0, 6)
        elif isinstance(v, str) and '%' in v:
            s = v.replace('%', '').strip()
            try:
                cell.value = float(s)
            except Exception:
                pass
        try:
            cell.number_format = '0.00'
        except Exception:
            pass
    except Exception:
        pass



@app.route('/', methods=['GET'])
def index():
    return render_template_string(HTML)


@app.route('/generate', methods=['POST'])
def generate():
    if 'mermas' not in request.files:
        flash('Falta fichero MERMAS')
        return redirect(url_for('index'))

    m = request.files['mermas']
    if not m or m.filename.strip() == '':
        flash('Falta fichero MERMAS')
        return redirect(url_for('index'))

    tmpdir = tempfile.mkdtemp()
    mer_path = os.path.join(tmpdir, 'mermas.xlsx')
    m.save(mer_path)

    tpl_art = os.path.join(os.path.dirname(__file__), 'PLANTILLA Artículos.xlsx')
    tpl_chk = os.path.join(os.path.dirname(__file__), 'CHECKLIST CALIDAD DE REPARTO.xlsx')

    if not os.path.exists(tpl_art) or not os.path.exists(tpl_chk):
        return "Faltan plantillas en la app. Contacta para añadirlas.", 500

    out_path = os.path.join(tmpdir, 'TOP_GENERADO.xlsx')

    try:
        generate_from_mermas(mer_path, tpl_art, tpl_chk, out_path, m.filename)
    except Exception as e:
        return f"Error durante generación: {e}", 500

    return send_file(out_path, as_attachment=True, download_name='TOP_GENERADO.xlsx')


def generate_from_mermas(mermas_path, tpl_art_path, tpl_chk_path, out_path, original_filename):
    wb_mermas = load_workbook(mermas_path, data_only=True)

    top_name = None
    for name in wb_mermas.sheetnames:
        if 'top' in name.lower() or 'calidad' in name.lower():
            top_name = name
            break
    if not top_name:
        top_name = wb_mermas.sheetnames[0]

    ws_top = wb_mermas[top_name]

    sem, cic = extract_sem_ciclo_from_name(original_filename)
    sem = sem or ''
    cic = cic or ''

    headers = [(ws_top.cell(row=1, column=c).value or "") for c in range(1, ws_top.max_column + 1)]
    low = [str(h).lower() for h in headers]

    def find_idx(keys):
        for k in keys:
            for i, h in enumerate(low):
                if k in h:
                    return i + 1
        return None

    mc_idx = find_idx(['mc', 'modelo', 'model', 'codigo', 'articulo']) or 1
    fam_idx = find_idx(['familia', 'family', 'grupo'])

    map_cols = [2, 3, 4, 5, 6, 7, 9, 10]

    mc_map = {}
    for r in range(2, ws_top.max_row + 1):
        val = ws_top.cell(row=r, column=mc_idx).value
        if val is None:
            continue

        key = str(val).strip()
        if not key:
            continue

        mapped = []
        for col in map_cols:
            mapped.append(ws_top.cell(row=r, column=col).value if col <= ws_top.max_column else None)

        fam = ws_top.cell(row=r, column=fam_idx).value if fam_idx and fam_idx <= ws_top.max_column else None
        mc_map[key] = {'mapped': mapped, 'fam': fam, 'row': r}

    wb_art = load_workbook(tpl_art_path, data_only=False)
    wb_chk = load_workbook(tpl_chk_path, data_only=False)

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    copy_sheet_exact(ws_top, wb_out, "TOP-CALIDAD-2")

    tpl_art_sheet = wb_art[wb_art.sheetnames[0]]
    tpl_chk_sheet = wb_chk[wb_chk.sheetnames[0]]

    existing = set(wb_out.sheetnames)
    pairs = []

    for key in mc_map:
        safe = sanitize_sheet_name(key)
        base = safe
        i = 1
        while safe in existing:
            suf = f'-{i}'
            safe = base[:31 - len(suf)] + suf
            i += 1

        existing.add(safe)

        new_mc = copy_sheet_exact(tpl_art_sheet, wb_out, safe)

        new_mc.cell(row=4, column=2).value = key
        if mc_map[key]['fam'] is not None:
            new_mc.cell(row=5, column=2).value = mc_map[key]['fam']

        try:
            new_mc.cell(row=2, column=2).value = int(sem) if sem != '' else None
        except Exception:
            new_mc.cell(row=2, column=2).value = sem

        try:
            new_mc.cell(row=2, column=4).value = int(cic) if cic != '' else None
        except Exception:
            new_mc.cell(row=2, column=4).value = cic

        for idx, val in enumerate(mc_map[key]['mapped'], start=1):
            new_mc.cell(row=48, column=idx).value = val
        new_mc.cell(row=48, column=9).value = None

        cell = new_mc.cell(row=48, column=6)
        convert_pct_cell_to_number(cell)

        chk_name = f'CHECKLIST-{safe}'
        chk = copy_sheet_exact(tpl_chk_sheet, wb_out, chk_name)

        a1 = chk.cell(row=1, column=1).value or ""
        if re.search(r"art[ií]culo\s*:", str(a1), flags=re.IGNORECASE):
            new_a1 = re.sub(
                r"(art[ií]culo\s*:\s*).*",
                lambda m: m.group(1) + str(key),
                str(a1),
                flags=re.IGNORECASE,
            )
        else:
            new_a1 = str(a1).strip() + " Artículo: " + str(key)
        chk.cell(row=1, column=1).value = new_a1

        chk.cell(row=1, column=4).value = f"SEMANA: {sem}   CICLO: {cic}"

        chk.cell(row=4, column=2).value = None
        chk.cell(row=5, column=2).value = None
        for c in range(1, 9):
            chk.cell(row=48, column=c).value = None

        pairs.append((safe, chk_name))

    ordered = ['TOP-CALIDAD-2']
    for a, b in pairs:
        ordered.append(a)
        ordered.append(b)

    for s in wb_out.sheetnames:
        if s not in ordered:
            ordered.append(s)

    wb_out._sheets = [wb_out[s] for s in ordered]
    wb_out.save(out_path)


def _extract_top_models_and_label(file_storage):
    file_bytes = BytesIO(file_storage.read())
    wb = load_workbook(file_bytes, data_only=True)

    top_name = None
    for name in wb.sheetnames:
        if 'top' in name.lower() or 'calidad' in name.lower():
            top_name = name
            break
    if not top_name:
        top_name = wb.sheetnames[0]

    ws = wb[top_name]

    headers = [(ws.cell(row=1, column=c).value or "") for c in range(1, ws.max_column + 1)]
    low = [str(h).lower() for h in headers]

    def find_idx(keys):
        for k in keys:
            for i, h in enumerate(low):
                if k in h:
                    return i + 1
        return None

    mc_idx = find_idx(['mc', 'modelo', 'model', 'codigo', 'articulo']) or 1
    fam_idx = find_idx(['familia', 'family', 'grupo'])

    sem, cic = extract_sem_ciclo_from_name(file_storage.filename)
    if not sem or not cic:
        top_sem = ws["B2"].value
        top_cic = ws["D2"].value
        if not sem:
            sem_match = re.search(r'\d+', str(top_sem) if top_sem is not None else "")
            if sem_match:
                sem = sem_match.group(0)
        if not cic:
            cic_match = re.search(r'\d+', str(top_cic) if top_cic is not None else "")
            if cic_match:
                cic = cic_match.group(0)

    label = f"SEM {sem} C{cic}" if sem or cic else file_storage.filename

    models = set()
    families = {}

    for r in range(2, ws.max_row + 1):
        mc = ws.cell(row=r, column=mc_idx).value
        if mc is None:
            continue
        mc = str(mc).strip()
        if not mc:
            continue
        models.add(mc)

        if fam_idx:
            fam = ws.cell(row=r, column=fam_idx).value
            if fam is not None and mc not in families:
                fam_txt = str(fam).strip()
                if fam_txt:
                    families[mc] = fam_txt

    return label, models, families


def _seed_records_from_master_workbook(master_wb):
    records = {}
    order = []

    if '_ORDEN' in master_wb.sheetnames:
        ws_order = master_wb['_ORDEN']
        for r in range(2, ws_order.max_row + 1):
            label = ws_order.cell(row=r, column=1).value
            if label is None:
                continue
            label = str(label).strip()
            if label:
                order.append(label)

    if 'DETALLE' not in master_wb.sheetnames:
        return records, order

    ws = master_wb['DETALLE']

    for r in range(2, ws.max_row + 1):
        mc = ws.cell(row=r, column=1).value
        if mc is None:
            continue
        mc = str(mc).strip()
        if not mc:
            continue

        count = ws.cell(row=r, column=2).value
        try:
            count = int(count)
        except Exception:
            count = 0

        ciclos_raw = ws.cell(row=r, column=3).value or ""
        labels = [c.strip() for c in str(ciclos_raw).split(",") if c.strip()]
        first_label = str(ws.cell(row=r, column=4).value or "").strip()
        last_label = str(ws.cell(row=r, column=5).value or "").strip()
        family = str(ws.cell(row=r, column=9).value or "").strip() if ws.max_column >= 9 else ""

        records[mc] = {
            "count": count if count else len(labels),
            "labels": labels,
            "first_label": first_label or (labels[0] if labels else "HISTORICO ACUMULADO"),
            "last_label": last_label or (labels[-1] if labels else "HISTORICO ACUMULADO"),
            "family": family,
        }

    return records, order


def _sequence_stats(order, labels):
    order_map = {label: idx for idx, label in enumerate(order)}
    positions = [order_map[label] for label in labels if label in order_map]
    positions = sorted(set(positions))
    if not positions:
        return 0

    longest = 1
    current = 1
    for i in range(1, len(positions)):
        if positions[i] == positions[i - 1] + 1:
            current += 1
            longest = max(longest, current)
        else:
            current = 1
    return longest


def _build_historico_rows(records, order, last_seen_models, last_label):
    rows = []
    family_totals = {}
    family_models = {}
    persistent_total_count = 0
    persistent_consecutive_count = 0

    for mc, rec in records.items():
        labels = rec.get("labels", [])
        family = rec.get("family") or "Sin familia"
        appears_in_last_top = mc in last_seen_models
        is_recurrent = len(labels) > 1

        racha_max = _sequence_stats(order, labels)
        persistent_total = len(labels) >= 3
        persistent_consecutive = racha_max >= 3

        if appears_in_last_top and is_recurrent:
            prioridad = "ALTA"
            estado = "Reincidente"
        elif appears_in_last_top:
            prioridad = "MEDIA"
            estado = "Nuevo"
        else:
            prioridad = "BAJA"
            estado = "Histórico"

        if persistent_total:
            persistent_total_count += 1
        if persistent_consecutive:
            persistent_consecutive_count += 1

        family_totals[family] = family_totals.get(family, 0) + len(labels)
        family_models.setdefault(family, set()).add(mc)

        rows.append({
            "mc": mc,
            "count": len(labels),
            "ciclos": ", ".join(labels),
            "first_label": rec.get("first_label", labels[0] if labels else ""),
            "last_label": rec.get("last_label", labels[-1] if labels else ""),
            "appears_in_last_top": "Sí" if appears_in_last_top else "No",
            "estado": estado,
            "prioridad": prioridad,
            "family": family,
            "persistencia_total": "Sí" if persistent_total else "No",
            "persistencia_consecutiva": "Sí" if persistent_consecutive else "No",
            "racha_max": racha_max,
        })

    priority_rank = {"ALTA": 0, "MEDIA": 1, "BAJA": 2}
    estado_rank = {"Reincidente": 0, "Nuevo": 1, "Histórico": 2}
    rows.sort(
        key=lambda r: (
            priority_rank.get(r["prioridad"], 99),
            estado_rank.get(r["estado"], 99),
            -r["count"],
            r["first_label"],
            r["mc"],
        )
    )

    total_models = len(rows)
    new_models = sum(1 for r in rows if r["estado"] == "Nuevo")
    repeated_models = sum(1 for r in rows if r["count"] > 1)

    family_rows = []
    for fam, total in family_totals.items():
        family_rows.append({
            "family": fam,
            "apariciones": total,
            "modelos_unicos": len(family_models.get(fam, set())),
        })
    family_rows.sort(key=lambda x: (-x["apariciones"], -x["modelos_unicos"], x["family"]))
    alerts = [r for r in rows if r["prioridad"] == "ALTA"][:10]

    return rows, total_models, new_models, repeated_models, persistent_total_count, persistent_consecutive_count, family_rows, alerts


def _render_historico_visual(rows, total_models, new_models, repeated_models, persistent_total_count, persistent_consecutive_count, family_rows, alerts, order, sheet_title="HISTORICO"):
    wb_out = Workbook()
    ws_res = wb_out.active
    ws_res.title = "RESUMEN"
    ws_res.sheet_view.showGridLines = False

    title_font = Font(bold=True, size=14, color="FFFFFF")
    header_font = Font(bold=True, color="FFFFFF")
    body_font = Font(bold=False, color="1F1F1F")
    bold_font = Font(bold=True, color="1F1F1F")

    fill_title = PatternFill("solid", fgColor="1F4E78")
    fill_header = PatternFill("solid", fgColor="4F81BD")
    fill_card = PatternFill("solid", fgColor="D9EAF7")
    fill_green = PatternFill("solid", fgColor="E2F0D9")
    fill_amber = PatternFill("solid", fgColor="FFF2CC")
    fill_red = PatternFill("solid", fgColor="F4CCCC")
    fill_gray = PatternFill("solid", fgColor="E7E6E6")
    fill_blue_soft = PatternFill("solid", fgColor="DDEBF7")

    thin = Side(style="thin", color="7F7F7F")
    medium = Side(style="medium", color="404040")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_medium = Border(left=medium, right=medium, top=medium, bottom=medium)

    ws_res.merge_cells("A1:E1")
    ws_res["A1"] = f"{sheet_title} - RESUMEN HISTÓRICO TOP MERMAS"
    ws_res["A1"].font = title_font
    ws_res["A1"].fill = fill_title
    ws_res["A1"].alignment = Alignment(horizontal="center", vertical="center")

    summary_cards = [
        ("Total de modelos", total_models, fill_card),
        ("Cuántos son nuevos", new_models, fill_green if new_models else fill_gray),
        ("Cuántos se repiten", repeated_models, fill_amber if repeated_models else fill_gray),
        ("Persistentes 3+ apariciones", persistent_total_count, fill_blue_soft if persistent_total_count else fill_gray),
        ("Persistentes 3+ seguidos", persistent_consecutive_count, fill_blue_soft if persistent_consecutive_count else fill_gray),
        ("Alertas ALTA", len(alerts), fill_red if alerts else fill_gray),
    ]

    for i, (label, value, fill) in enumerate(summary_cards, start=3):
        ws_res.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
        ws_res.merge_cells(start_row=i, start_column=3, end_row=i, end_column=5)
        c1 = ws_res.cell(row=i, column=1, value=label)
        c2 = ws_res.cell(row=i, column=3, value=value)
        c1.font = bold_font
        c2.font = bold_font
        c1.fill = fill
        c2.fill = fill
        c1.border = border_medium
        c2.border = border_medium
        c1.alignment = Alignment(horizontal="left", vertical="center")
        c2.alignment = Alignment(horizontal="center", vertical="center")

    alert_start = 11
    ws_res.merge_cells(start_row=alert_start, start_column=1, end_row=alert_start, end_column=5)
    ws_res.cell(row=alert_start, column=1, value="ALERTAS AUTOMÁTICAS (PRIORIDAD ALTA)").font = Font(bold=True, color="FFFFFF")
    ws_res.cell(row=alert_start, column=1).fill = fill_title
    ws_res.cell(row=alert_start, column=1).alignment = Alignment(horizontal="center", vertical="center")

    alert_headers = ["MC", "Familia", "Veces", "Estado", "Prioridad"]
    for col, h in enumerate(alert_headers, start=1):
        cell = ws_res.cell(row=alert_start + 1, column=col, value=h)
        cell.font = header_font
        cell.fill = fill_header
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="center", vertical="center")

    if alerts:
        for i, r in enumerate(alerts, start=alert_start + 2):
            vals = [r["mc"], r["family"], r["count"], r["estado"], r["prioridad"]]
            for col, v in enumerate(vals, start=1):
                cell = ws_res.cell(row=i, column=col, value=v)
                cell.border = border_thin
                cell.font = body_font
                cell.alignment = Alignment(horizontal="left" if col in (1, 2, 4, 5) else "center", vertical="center")
                cell.fill = fill_red
    else:
        ws_res.merge_cells(start_row=alert_start + 2, start_column=1, end_row=alert_start + 2, end_column=5)
        c = ws_res.cell(row=alert_start + 2, column=1, value="No hay modelos con prioridad ALTA.")
        c.border = border_thin
        c.fill = fill_gray
        c.alignment = Alignment(horizontal="center", vertical="center")

    family_start = alert_start + 7
    ws_res.merge_cells(start_row=family_start, start_column=1, end_row=family_start, end_column=5)
    ws_res.cell(row=family_start, column=1, value="FAMILIAS CON MÁS MERMAS").font = Font(bold=True, color="FFFFFF")
    ws_res.cell(row=family_start, column=1).fill = fill_title
    ws_res.cell(row=family_start, column=1).alignment = Alignment(horizontal="center", vertical="center")

    fam_headers = ["Familia", "Apariciones", "Modelos únicos"]
    for col, h in enumerate(fam_headers, start=1):
        cell = ws_res.cell(row=family_start + 1, column=col, value=h)
        cell.font = header_font
        cell.fill = fill_header
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, fam in enumerate(family_rows[:10], start=family_start + 2):
        vals = [fam["family"], fam["apariciones"], fam["modelos_unicos"]]
        for col, v in enumerate(vals, start=1):
            cell = ws_res.cell(row=i, column=col, value=v)
            cell.border = border_thin
            cell.font = body_font
            cell.alignment = Alignment(horizontal="left" if col == 1 else "center", vertical="center")
            cell.fill = fill_blue_soft if i % 2 == 0 else fill_card

    top_start = family_start + 14
    ws_res.merge_cells(start_row=top_start, start_column=1, end_row=top_start, end_column=6)
    ws_res.cell(row=top_start, column=1, value="TOP 10 REINCIDENTES").font = Font(bold=True, color="FFFFFF")
    ws_res.cell(row=top_start, column=1).fill = fill_title
    ws_res.cell(row=top_start, column=1).alignment = Alignment(horizontal="center", vertical="center")

    res_headers = ["MC", "Familia", "Veces", "Ciclos", "Estado", "Prioridad"]
    for col, h in enumerate(res_headers, start=1):
        cell = ws_res.cell(row=top_start + 1, column=col, value=h)
        cell.font = header_font
        cell.fill = fill_header
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="center", vertical="center")

    top_10 = sorted([r for r in rows if r["count"] > 1], key=lambda x: (-x["count"], x["mc"]))[:10]
    for i, r in enumerate(top_10, start=top_start + 2):
        vals = [r["mc"], r["family"], r["count"], r["ciclos"], r["estado"], r["prioridad"]]
        for col, v in enumerate(vals, start=1):
            cell = ws_res.cell(row=i, column=col, value=v)
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left" if col in (1, 2, 4, 5, 6) else "center", vertical="center")
            cell.font = body_font
            if r["prioridad"] == "ALTA":
                cell.fill = fill_red
            elif r["prioridad"] == "MEDIA":
                cell.fill = fill_amber
            else:
                cell.fill = fill_green if r["estado"] == "Nuevo" else fill_gray

    for col, width in {1: 24, 2: 18, 3: 12, 4: 56, 5: 16, 6: 14}.items():
        ws_res.column_dimensions[get_column_letter(col)].width = width

    ws_res.freeze_panes = "A3"

    ws_det = wb_out.create_sheet("DETALLE")
    ws_det.sheet_view.showGridLines = False

    det_headers = [
        "MC",
        "Veces",
        "Ciclos",
        "Primera aparición",
        "Última aparición",
        "¿Aparece en el último TOP?",
        "Estado",
        "Prioridad",
        "Familia",
        "Persistencia 3+ apariciones",
        "Persistencia 3+ seguidos",
        "Racha máx.",
    ]
    for col, h in enumerate(det_headers, start=1):
        cell = ws_det.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = fill_header
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, r in enumerate(rows, start=2):
        vals = [
            r["mc"], r["count"], r["ciclos"], r["first_label"], r["last_label"],
            r["appears_in_last_top"], r["estado"], r["prioridad"], r["family"],
            r["persistencia_total"], r["persistencia_consecutiva"], r["racha_max"],
        ]
        for col, v in enumerate(vals, start=1):
            cell = ws_det.cell(row=row_idx, column=col, value=v)
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left" if col not in (2, 12) else "center", vertical="center")
            cell.font = body_font
            if r["prioridad"] == "ALTA":
                cell.fill = fill_red
            elif r["prioridad"] == "MEDIA":
                cell.fill = fill_amber
            else:
                cell.fill = fill_green if r["estado"] == "Nuevo" else fill_gray

    widths = {1: 18, 2: 10, 3: 44, 4: 18, 5: 18, 6: 24, 7: 16, 8: 14, 9: 18, 10: 20, 11: 20, 12: 12}
    for col, width in widths.items():
        ws_det.column_dimensions[get_column_letter(col)].width = width

    ws_det.freeze_panes = "A2"
    ws_det.auto_filter.ref = ws_det.dimensions

    ws_ord = wb_out.create_sheet("_ORDEN")
    ws_ord.sheet_state = "hidden"
    ws_ord["A1"] = "Etiqueta"
    for i, label in enumerate(order, start=2):
        ws_ord.cell(row=i, column=1, value=label)

    output = BytesIO()
    wb_out.save(output)
    output.seek(0)
    return output





def _render_dashboard_html(rows, total_models, new_models, repeated_models, persistent_total_count, persistent_consecutive_count, family_rows, alerts, title="DASHBOARD TOP MERMAS"):
    def esc(v):
        import html
        return html.escape("" if v is None else str(v))

    def badge(text, kind):
        colors = {
            "ALTA": "#f8d7da",
            "MEDIA": "#fff3cd",
            "BAJA": "#d1ecf1",
            "Reincidente": "#f8d7da",
            "Nuevo": "#d4edda",
            "Histórico": "#e2e3e5",
            "Sí": "#d4edda",
            "No": "#f8d7da",
        }
        return f'<span style="padding:4px 10px;border-radius:999px;background:{colors.get(kind, "#e2e3e5")};display:inline-block;">{esc(text)}</span>'

    def evo_text(ciclos_text):
        ciclos = [c.strip() for c in str(ciclos_text or "").split(",") if c.strip()]
        return " → ".join(ciclos) if ciclos else "-"

    cards = [
        ("Total de modelos", total_models, "#d9eaf7"),
        ("Nuevos", new_models, "#d4edda"),
        ("Repetidos", repeated_models, "#fff3cd"),
        ("Persistentes 3+ apariciones", persistent_total_count, "#d9edf7"),
        ("Persistentes 3+ seguidos", persistent_consecutive_count, "#d9edf7"),
        ("Alertas ALTA", len(alerts), "#f8d7da"),
    ]

    html = f"""
<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <title>{esc(title)}</title>
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <style>
    body {{
      font-family: Arial, Helvetica, sans-serif;
      margin: 0;
      background: #f6f8fb;
      color: #1f2937;
    }}
    .wrap {{
      max-width: 1400px;
      margin: 0 auto;
      padding: 24px;
    }}
    h1, h2, h3 {{
      margin: 0 0 12px 0;
    }}
    .sub {{
      color: #6b7280;
      margin-bottom: 18px;
    }}
    .grid {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
      gap: 14px;
      margin: 18px 0 28px 0;
    }}
    .card {{
      background: white;
      border: 1px solid #e5e7eb;
      border-radius: 16px;
      padding: 16px;
      box-shadow: 0 6px 20px rgba(15, 23, 42, 0.04);
    }}
    .card .label {{
      font-size: 13px;
      color: #6b7280;
      margin-bottom: 8px;
    }}
    .card .value {{
      font-size: 30px;
      font-weight: 700;
      line-height: 1;
    }}
    .section {{
      background: white;
      border: 1px solid #e5e7eb;
      border-radius: 16px;
      padding: 16px;
      box-shadow: 0 6px 20px rgba(15, 23, 42, 0.04);
      margin-bottom: 18px;
    }}
    .table-wrap {{
      overflow: auto;
      border-radius: 12px;
      border: 1px solid #e5e7eb;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      background: white;
      min-width: 900px;
    }}
    th, td {{
      border-bottom: 1px solid #e5e7eb;
      padding: 10px 12px;
      text-align: left;
      vertical-align: top;
      font-size: 14px;
    }}
    th {{
      position: sticky;
      top: 0;
      background: #eef4fb;
      cursor: pointer;
      user-select: none;
      z-index: 1;
    }}
    tr:hover td {{
      background: #fafcff;
    }}
    .filters {{
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
      margin-bottom: 12px;
    }}
    .filters input {{
      padding: 10px 12px;
      border-radius: 10px;
      border: 1px solid #d1d5db;
      min-width: 240px;
    }}
    .small {{
      font-size: 13px;
      color: #6b7280;
    }}
    .row-red td {{
      background: #fff5f5;
    }}
    .row-amber td {{
      background: #fffdf0;
    }}
    .row-green td {{
      background: #f7fff7;
    }}
    .detail-row td {{
      background: #f8fafc;
      border-bottom: 1px solid #dbe3ec;
      font-size: 13px;
      color: #374151;
      padding-top: 12px;
      padding-bottom: 12px;
    }}
    .btn {{
      display: inline-block;
      padding: 10px 14px;
      border-radius: 10px;
      text-decoration: none;
      background: #1f4e78;
      color: white;
      font-weight: 700;
      margin-bottom: 14px;
    }}
    .clickable-mc {{
      cursor: pointer;
      font-weight: 700;
      color: #1f4e78;
    }}
    .hint {{
      font-size: 12px;
      color: #6b7280;
      margin-top: 8px;
    }}
  </style>
</head>
<body>
  <div class="wrap">
    <a class="btn" href="/">Volver</a>
    <h1>{esc(title)}</h1>
    <div class="sub">Panel interactivo de histórico: búsqueda, ordenación y drill-down rápido sin descargar Excel.</div>

    <div class="grid">
"""
    for label, value, color in cards:
        html += f"""
      <div class="card" style="background:{color};">
        <div class="label">{esc(label)}</div>
        <div class="value">{esc(value)}</div>
      </div>"""
    html += """
    </div>

    <div class="section">
      <h2>Alertas automáticas</h2>
      <div class="small">Modelos con prioridad ALTA</div>
      <div class="table-wrap">
        <table id="table-alerts">
          <thead>
            <tr>
              <th>MC</th><th>Familia</th><th>Veces</th><th>Estado</th><th>Prioridad</th>
            </tr>
          </thead>
          <tbody>
"""
    if alerts:
        for r in alerts:
            html += f"""
            <tr class="row-red">
              <td>{esc(r.get('mc'))}</td>
              <td>{esc(r.get('family'))}</td>
              <td>{esc(r.get('count'))}</td>
              <td>{badge(r.get('estado'), r.get('estado'))}</td>
              <td>{badge(r.get('prioridad'), r.get('prioridad'))}</td>
            </tr>"""
    else:
        html += """
            <tr><td colspan="5">No hay alertas ALTA.</td></tr>"""
    html += """
          </tbody>
        </table>
      </div>
    </div>

    <div class="section">
      <h2>Familias con más mermas</h2>
      <div class="table-wrap">
        <table id="table-family">
          <thead>
            <tr><th>Familia</th><th>Apariciones</th><th>Modelos únicos</th></tr>
          </thead>
          <tbody>
"""
    for fam in family_rows:
        html += f"""
            <tr>
              <td>{esc(fam.get('family'))}</td>
              <td>{esc(fam.get('apariciones'))}</td>
              <td>{esc(fam.get('modelos_unicos'))}</td>
            </tr>"""
    html += """
          </tbody>
        </table>
      </div>
    </div>

    <div class="section">
      <h2>Detalle completo</h2>
      <div class="filters">
        <input id="searchInput" type="text" placeholder="Buscar MC, familia, estado...">
      </div>
      <div class="table-wrap">
        <table id="table-detail">
          <thead>
            <tr>
              <th data-sort="text">MC</th>
              <th data-sort="num">Veces</th>
              <th data-sort="text">Ciclos</th>
              <th data-sort="text">Primera aparición</th>
              <th data-sort="text">Última aparición</th>
              <th data-sort="text">Último TOP</th>
              <th data-sort="text">Estado</th>
              <th data-sort="text">Prioridad</th>
              <th data-sort="text">Familia</th>
              <th data-sort="text">Persist. 3+</th>
              <th data-sort="text">Persist. seg.</th>
              <th data-sort="num">Racha máx.</th>
            </tr>
          </thead>
          <tbody>
"""
    for r in rows:
        cls = "row-red" if r.get("prioridad") == "ALTA" else ("row-amber" if r.get("prioridad") == "MEDIA" else ("row-green" if r.get("estado") == "Nuevo" else ""))
        ciclos_text = r.get('ciclos') or ""
        evo = evo_text(ciclos_text)
        html += f"""
            <tr class="{cls}" onclick="toggleDetails(this)" style="cursor:pointer;">
              <td class="clickable-mc" title="Haz clic para ver el detalle">{esc(r.get('mc'))}</td>
              <td>{esc(r.get('count'))}</td>
              <td>{esc(ciclos_text)}</td>
              <td>{esc(r.get('first_label'))}</td>
              <td>{esc(r.get('last_label'))}</td>
              <td>{badge(r.get('appears_in_last_top'), r.get('appears_in_last_top'))}</td>
              <td>{badge(r.get('estado'), r.get('estado'))}</td>
              <td>{badge(r.get('prioridad'), r.get('prioridad'))}</td>
              <td>{esc(r.get('family'))}</td>
              <td>{badge(r.get('persistencia_total'), r.get('persistencia_total'))}</td>
              <td>{badge(r.get('persistencia_consecutiva'), r.get('persistencia_consecutiva'))}</td>
              <td>{esc(r.get('racha_max'))}</td>
            </tr>
            <tr class="detail-row" style="display:none;">
              <td colspan="12">
                <b>Ciclos:</b> {esc(ciclos_text)}<br>
                <b>Evolución:</b> {esc(evo)}
              </td>
            </tr>"""
    html += """
          </tbody>
        </table>
      </div>
      <div class="hint">Haz clic sobre un MC para desplegar su evolución.</div>
    </div>
  </div>

  <script>
    const searchInput = document.getElementById('searchInput');
    const detailTable = document.getElementById('table-detail');

    searchInput.addEventListener('input', () => {
      const q = searchInput.value.toLowerCase().trim();
      const allRows = Array.from(detailTable.querySelectorAll('tbody tr'));
      for (let i = 0; i < allRows.length; i += 2) {
        const mainRow = allRows[i];
        const extraRow = allRows[i + 1];
        const txt = (mainRow.innerText + ' ' + extraRow.innerText).toLowerCase();
        const show = txt.includes(q);
        mainRow.style.display = show ? '' : 'none';
        if (!show) {
          extraRow.style.display = 'none';
          extraRow.dataset.open = '0';
        } else {
          extraRow.style.display = extraRow.dataset.open === '1' ? 'table-row' : 'none';
        }
      }
    });

    function sortTable(table, colIndex, type) {
      const tbody = table.tBodies[0];
      const allRows = Array.from(tbody.querySelectorAll('tr'));
      const pairs = [];
      for (let i = 0; i < allRows.length; i += 2) {
        pairs.push([allRows[i], allRows[i + 1]]);
      }
      const asc = table.getAttribute('data-sort-dir') !== 'asc';
      pairs.sort((a, b) => {
        let av = a[0].children[colIndex].innerText.trim();
        let bv = b[0].children[colIndex].innerText.trim();
        if (type === 'num') {
          av = parseFloat(av.replace(',', '.')) || 0;
          bv = parseFloat(bv.replace(',', '.')) || 0;
          return asc ? av - bv : bv - av;
        }
        return asc ? av.localeCompare(bv, 'es') : bv.localeCompare(av, 'es');
      });
      pairs.forEach(([main, extra]) => {
        tbody.appendChild(main);
        tbody.appendChild(extra);
      });
      table.setAttribute('data-sort-dir', asc ? 'asc' : 'desc');
    }

    function toggleDetails(row) {
      const extra = row.nextElementSibling;
      if (!extra || !extra.classList.contains('detail-row')) return;
      const isOpen = extra.style.display === 'table-row';
      extra.style.display = isOpen ? 'none' : 'table-row';
      extra.dataset.open = isOpen ? '0' : '1';
    }

    document.querySelectorAll('#table-detail thead th').forEach((th, idx) => {
      th.addEventListener('click', () => sortTable(detailTable, idx, th.dataset.sort || 'text'));
    });
  </script>
</body>
</html>
"""
    return html


@app.route('/dashboard', methods=['POST'])
def dashboard():
    uploaded_files = request.files.getlist('files')
    valid_files = [f for f in uploaded_files if f and f.filename and f.filename.lower().endswith(('.xlsx', '.xlsm', '.xls'))]
    if not valid_files:
        return "No has subido ficheros Excel válidos para el dashboard", 400

    records = {}
    order = []
    last_seen_models = set()
    last_label = None

    for f in valid_files:
        try:
            label, models, families = _extract_top_models_and_label(f)
            order.append(label)
            last_label = label
            last_seen_models = models
            for mc in models:
                if mc not in records:
                    records[mc] = {
                        "count": 0,
                        "labels": [],
                        "first_label": label,
                        "last_label": label,
                        "family": families.get(mc, ""),
                    }
                rec = records[mc]
                rec["count"] += 1
                if label not in rec["labels"]:
                    rec["labels"].append(label)
                if not rec.get("family") and families.get(mc):
                    rec["family"] = families.get(mc)
                rec["last_label"] = label
        except Exception:
            continue

    if not records:
        return "No se han podido leer datos válidos para el dashboard", 400

    rows, total_models, new_models, repeated_models, persistent_total_count, persistent_consecutive_count, family_rows, alerts = _build_historico_rows(records, order, last_seen_models, last_label)
    return _render_dashboard_html(rows, total_models, new_models, repeated_models, persistent_total_count, persistent_consecutive_count, family_rows, alerts, title="DASHBOARD TOP MERMAS")
@app.route('/historico_acumulado', methods=['POST'])
def generar_historico_acumulado():
    master_file = request.files.get('master')
    uploaded_files = request.files.getlist('files')

    valid_files = [f for f in uploaded_files if f and f.filename and f.filename.lower().endswith(('.xlsx', '.xlsm', '.xls'))]
    if not valid_files:
        return "No has subido ficheros Excel válidos para actualizar el histórico acumulado", 400

    records = {}
    order = []

    if master_file and master_file.filename and master_file.filename.lower().endswith('.xlsx'):
        try:
            master_bytes = BytesIO(master_file.read())
            master_wb = load_workbook(master_bytes, data_only=True)
            records, order = _seed_records_from_master_workbook(master_wb)
        except Exception:
            records = {}
            order = []

    last_seen_models = set()
    last_label = None

    for f in valid_files:
        try:
            label, models, families = _extract_top_models_and_label(f)
            order.append(label)
            last_label = label
            last_seen_models = models

            for mc in models:
                if mc not in records:
                    records[mc] = {
                        "count": 0,
                        "labels": [],
                        "first_label": label,
                        "last_label": label,
                        "family": families.get(mc, ""),
                    }
                rec = records[mc]
                rec["count"] += 1
                if label not in rec["labels"]:
                    rec["labels"].append(label)
                if not rec.get("family") and families.get(mc):
                    rec["family"] = families.get(mc)
                rec["last_label"] = label

        except Exception:
            continue

    if not records:
        return "No se han podido consolidar datos para el histórico acumulado", 400

    rows, total_models, new_models, repeated_models, persistent_total_count, persistent_consecutive_count, family_rows, alerts = _build_historico_rows(records, order, last_seen_models, last_label)
    output = _render_historico_visual(rows, total_models, new_models, repeated_models, persistent_total_count, persistent_consecutive_count, family_rows, alerts, order, sheet_title="HISTORICO ACUMULADO")

    return send_file(
        output,
        as_attachment=True,
        download_name="historico_acumulado.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
